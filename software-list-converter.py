#!/usr/bin/env python3
from openpyxl import load_workbook, Workbook
import argparse
import os
COLUMN = 'C'

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Convert application/version format in a .xlsx file')
    parser.add_argument('input_file', help='Input .xlsx file')
    args = parser.parse_args()
    wb = load_workbook(args.input_file)
    wb2 = Workbook()
    ws = wb.active
    ws2 = wb2.active
    ws2.title = 'Full - converted'
    ws2.column_dimensions['A'].width = 27
    ws2.column_dimensions['B'].width = 17
    ws2.column_dimensions['C'].width = 54
    ws2.column_dimensions['D'].width = 24

    print('Converting file ...')
    for i, row in enumerate(ws.rows):
        if i % 100 == 0 and i > 0:
            print('{} rows converted...'.format(i))
        if i < 4:
            ws2.append([cell.value for cell in row])
        else:
            index = '{}{}'.format(COLUMN, i+1)
            value = row[2].value
            if ',' in value:
                arr = value.split(',')
                for j in range(len(arr)):
                    cell = arr[j]
                    if '|' in cell:
                        data = cell.split('|')
                        software = '|'.join(data[:-1])
                        version = data[-1]
                        ws2.append([row[0].value, row[1].value, software, version])
                        if (len(data) > 2):
                            pass
                    else:
                        # print('{}: "{}"'.format(index, cell))
                        ws2.append([row[0].value, row[1].value, cell])
            else:
                # print('{}: ["{}": "{}": "{}"]'.format(index, row[0].value, row[1].value, row[2].value))
                ws2.append([cell.value for cell in row])

    print('{} rows converted.'.format(ws.max_row))
    ws2['B2'] = '4 columns'
    ws2['B3'] = '{} rows'.format(ws2.max_row - 4)
    ws2['C4'] = 'Application'
    ws2['D4'] = 'Version'

    base, ext = os.path.splitext(args.input_file)
    output_name = '{}-converted{}'.format(base, ext)
    print('Saving converted file to "{}"'.format(output_name))
    wb2.save(output_name)
